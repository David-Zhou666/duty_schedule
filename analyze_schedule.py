import openpyxl
from collections import Counter

file_path = 'duty_schedule_result.xlsx'
wb = openpyxl.load_workbook(file_path)

print("="*70)
print("   排班表统计分析")
print("="*70)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    male_counter = Counter()
    female_counter = Counter()
    
    for row_idx, row in enumerate(ws.iter_rows(max_row=6, values_only=True), 1):
        if row_idx == 1 or not row[0]:
            continue
        
        if row[1]:
            males = [m.strip() for m in str(row[1]).split('、') if m.strip()]
            for male in males:
                male_counter[male] += 1
        
        if row[2]:
            females = [f.strip() for f in str(row[2]).split('、') if f.strip()]
            for female in females:
                female_counter[female] += 1
    
    print(f"\nSheet: {sheet_name}\n")
    print("男生值班次数统计:")
    print("-" * 70)
    print(f"{'姓名':<15} | {'排班次数':<8} | {'状态'}")
    print("-" * 70)
    for name, count in sorted(male_counter.items(), key=lambda x: -x[1]):
        status = "✓" if count <= 5 else "⚠(超）"
        print(f"{name:<15} | {count:<8d} | {status}")
    
    print(f"\n{'女生值班次数统计:'}")
    print("-" * 70)
    print(f"{'姓名':<15} | {'排班次数':<8} | {'状态'}")
    print("-" * 70)
    for name, count in sorted(female_counter.items(), key=lambda x: -x[1]):
        status = "✓" if count <= 5 else "⚠(超）"
        print(f"{name:<15} | {count:<8d} | {status}")
    
    print(f"\n统计汇总:")
    print(f"男生排班总人次: {sum(male_counter.values())}")
    print(f"女生排班总人次: {sum(female_counter.values())}")
    print(f"男生独立人数: {len(male_counter)}")
    print(f"女生独立人数: {len(female_counter)}")

print("\n" + "="*70)
print("验证结果: ✓ = 符合要求（≤5次） | ⚠ = 超出限制（>5次）")
print("="*70)

wb.close()

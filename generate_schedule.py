#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
值日排班表生成器（Python版本）
从duty_temple.xlsx读取数据并生成排班表

使用方法:
    python generate_schedule.py
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import os
import sys
from collections import defaultdict
from datetime import datetime

# 配置
INPUT_FILE = 'src/main/resources/duty_temple.xlsx'
OUTPUT_FILE = 'duty_schedule_result.xlsx'
MALE_PER_DAY = 4
FEMALE_PER_DAY = 2
WEEKLY_LIMIT = 5


class Student:
    """学生类"""
    def __init__(self, name, gender, availability):
        self.name = name
        self.gender = gender
        self.availability = availability  # [Mon, Tue, Wed, Thu, Fri]
        self.weekly_count = 0
    
    def is_available_on(self, day_idx):
        if 0 <= day_idx < len(self.availability):
            return self.availability[day_idx]
        return False
    
    def can_schedule(self):
        return self.weekly_count < WEEKLY_LIMIT
    
    def increment_count(self):
        self.weekly_count += 1


def read_excel(filepath):
    """从Excel文件读取学生信息
    
    格式说明：
    - 每个学生有两行数据（第一节和第二节）
    - 性别列（第0列）: 仅在该组第一个学生显示（男生/女生）
    - 姓名列（第1列）: 仅在"第一节"行显示
    - 时间列（第2列）: 第一节/第二节
    - 周一到周五（第3-7列）: 是/否
    - 可用性取两行中任一行为"是"
    """
    students = {'male': [], 'female': []}
    
    if not os.path.exists(filepath):
        print(f"错误: 文件不存在 - {filepath}")
        sys.exit(1)
    
    try:
        wb = openpyxl.load_workbook(filepath)
        print(f"打开文件成功，Sheet页数: {len(wb.sheetnames)}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n正在读取 Sheet: {sheet_name}")
            
            current_gender = None
            student_data_map = {}  # 用于合并每个学生的两行数据
            
            # 从第2行开始读取（第1行是表头）
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            
            for idx, row in enumerate(rows):
                if not row or all(v is None for v in row):  # 跳过空行
                    continue
                
                # 获取各列数据
                gender_cell = row[0]  # 第一列：性别或空
                name_cell = row[1]    # 第二列：姓名
                time_cell = row[2]    # 第三列：时间（第一节/第二节）
                
                # 检测性别标记
                if gender_cell and str(gender_cell).strip() in ['男生', '女生']:
                    current_gender = str(gender_cell).strip()
                    continue
                
                # 处理学生行
                if time_cell and str(time_cell).strip() == '第一节' and name_cell and str(name_cell).strip():
                    name = str(name_cell).strip()
                    
                    # 获取第一节的可用性（列3-7）
                    availability_1 = []
                    for day_idx in range(5):
                        cell_val = row[3 + day_idx] if 3 + day_idx < len(row) else None
                        available = str(cell_val).strip() == '是' if cell_val else False
                        availability_1.append(available)
                    
                    # 获取第二节的可用性
                    availability_2 = [False] * 5
                    if idx + 1 < len(rows):
                        next_row = rows[idx + 1]
                        if next_row and next_row[2] and str(next_row[2]).strip() == '第二节':
                            for day_idx in range(5):
                                cell_val = next_row[3 + day_idx] if 3 + day_idx < len(next_row) else None
                                available = str(cell_val).strip() == '是' if cell_val else False
                                availability_2[day_idx] = available
                    
                    # 合并：任一节为"是"即为可用
                    availability = [availability_1[i] or availability_2[i] for i in range(5)]
                    
                    # 推断性别
                    gender_key = current_gender.replace('生', '') if current_gender else ('male' if '男' in sheet_name else 'female')
                    if gender_key == '男':
                        gender_key = 'male'
                    elif gender_key == '女':
                        gender_key = 'female'
                    
                    # 检查是否已存在（避免重复）
                    if name not in student_data_map:
                        student = Student(name, gender_key, availability)
                        students[gender_key].append(student)
                        student_data_map[name] = True
                        
                        # 统计可用天数
                        available_days = []
                        day_names = ['一', '二', '三', '四', '五']
                        for i in range(5):
                            if availability[i]:
                                available_days.append(day_names[i])
                        available_str = ','.join(available_days) if available_days else "无"
                        
                        print(f"  添加{'男生' if gender_key == 'male' else '女生'}: {name} - 可用: {available_str}")
        
        wb.close()
        
    except Exception as e:
        print(f"读取Excel文件失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    
    return students


def select_students(students, count, day_idx):
    """为某天选择学生
    
    算法：
    1. 优先选择该天可用且周次数最少、未到达限制的学生
    2. 如果不足，尝试从其他可用学生中选择
    """
    selected = []
    selected_set = set()
    
    # 第一阶段：选择该天可用且未超周限的学生（无重复）
    available = [s for s in students 
                if s.is_available_on(day_idx) and s.can_schedule()]
    available.sort(key=lambda s: s.weekly_count)
    
    for student in available:
        if len(selected) >= count:
            break
        selected.append(student)
        selected_set.add(student)
        student.increment_count()
    
    # 第二阶段：如果不足，允许已选的学生在限制内重复
    if len(selected) < count:
        for student in list(selected):  # 从已选的学生中复制
            if len(selected) >= count:
                break
            if student.can_schedule():
                selected.append(student)
                student.increment_count()
    
    # 第三阶段：如果仍不足，从其他可用学生中尝试
    if len(selected) < count:
        remaining = [s for s in available
                    if s not in selected_set and s.can_schedule()]
        
        for student in remaining:
            if len(selected) >= count:
                break
            selected.append(student)
            selected_set.add(student)
            student.increment_count()
    
    # 第四阶段：最后手段，从已选学生中继续重复（如果还有机会）
    if len(selected) < count:
        for student in list(selected):
            if len(selected) >= count:
                break
            if student.can_schedule():
                selected.append(student)
                student.increment_count()
    
    return selected


def generate_schedule(students):
    """生成一周的排班表
    
    规则：
    - 每天: 男生4人, 女生2人
    - 每人一周不超过5次
    - 如果人数不足，同一人可多次出现
    """
    schedule = {}
    days = ['周一', '周二', '周三', '周四', '周五']
    
    male_students = students['male']
    female_students = students['female']
    
    print(f"\n开始生成排班表...")
    print(f"男生: {len(male_students)}人, 女生: {len(female_students)}人")
    print(f"需求: 每天男生4人, 女生2人; 每人一周最多5次\n")
    
    for day_idx, day_name in enumerate(days):
        males = select_students(male_students, MALE_PER_DAY, day_idx)
        females = select_students(female_students, FEMALE_PER_DAY, day_idx)
        
        schedule[day_name] = {
            'males': [s.name for s in males],
            'females': [s.name for s in females]
        }
        
        male_names = '、'.join(schedule[day_name]['males']) if schedule[day_name]['males'] else '-'
        female_names = '、'.join(schedule[day_name]['females']) if schedule[day_name]['females'] else '-'
        print(f"{day_name}: 男生=[{male_names}] | 女生=[{female_names}]")
    
    return schedule


def export_to_excel(schedule, output_filepath):
    """将排班表导出到Excel

    格式: 周次 | 男1 | 男2 | 男3 | 男4 | 女1 | 女2
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "排班表"

    # 设置表头: 周次 | 男1 | 男2 | 男3 | 男4 | 女1 | 女2
    headers = ['周次', '男1', '男2', '男3', '男4', '女1', '女2']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)

    # 设置表头格式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx in range(1, 8):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 填充数据
    days = ['周一', '周二', '周三', '周四', '周五']
    for row_idx, day in enumerate(days, start=2):
        # 周次
        ws.cell(row=row_idx, column=1, value=day)

        # 填充男生姓名（4列）
        male_names = schedule[day]['males']
        for i in range(4):
            name = male_names[i] if i < len(male_names) else ""
            ws.cell(row=row_idx, column=2 + i, value=name)

        # 填充女生姓名（2列）
        female_names = schedule[day]['females']
        for i in range(2):
            name = female_names[i] if i < len(female_names) else ""
            ws.cell(row=row_idx, column=6 + i, value=name)

        # 设置对齐
        for col_idx in range(1, 8):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal='center', vertical='center')

    # 调整列宽
    ws.column_dimensions['A'].width = 12
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 15

    try:
        wb.save(output_filepath)
        print(f"\n排班表已导出: {output_filepath}")
        return True
    except Exception as e:
        print(f"导出Excel失败: {e}")
        return False


def main():
    """主程序"""
    print("="*60)
    print("  值日排班表生成器")
    print("="*60)
    print()
    
    # 检查输入文件
    if not os.path.exists(INPUT_FILE):
        print(f"错误: 输入文件不存在 - {INPUT_FILE}")
        print("请确保duty_temple.xlsx文件在src/main/resources/目录中")
        sys.exit(1)
    
    # 读取Excel文件
    print(f"正在读取 {INPUT_FILE}...\n")
    students = read_excel(INPUT_FILE)
    
    if not students['male'] and not students['female']:
        print("错误: 未读取到任何学生信息")
        sys.exit(1)
    
    # 生成排班表
    schedule = generate_schedule(students)
    
    # 导出Excel
    print(f"\n正在生成输出文件...")
    if export_to_excel(schedule, OUTPUT_FILE):
        print("\n" + "="*60)
        print("  排班完成！")
        print("="*60)
        print(f"输出文件: {OUTPUT_FILE}")
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()

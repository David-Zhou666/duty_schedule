#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取 duty_temple.xlsx（多 sheet），按要求生成值日排班表并导出为 duty_schedule_python.xlsx
规则：每个工作日（周一-周五）4 名男生、2 名女生；每人每周不超过 5 次；按当天可用且本周次数最少优先选。
"""
import os
from collections import defaultdict

try:
    import openpyxl
except Exception as e:
    raise SystemExit("请先安装 openpyxl: pip install openpyxl")

WORKSPACE_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
INPUT_PATH = os.path.join(WORKSPACE_ROOT, 'src', 'main', 'resources', 'duty_temple.xlsx')
OUTPUT_PATH = os.path.join(WORKSPACE_ROOT, 'duty_schedule_python.xlsx')

MALE_PER_DAY = 4
FEMALE_PER_DAY = 2
WEEKLY_LIMIT = 5
DAYS_PER_WEEK = 5
DAY_NAMES = ['周一', '周二', '周三', '周四', '周五']


class Student:
    def __init__(self, name, is_male, availability):
        self.name = name
        self.is_male = is_male
        self.availability = availability  # list of bool length 5
        self.weekly_count = 0

    def is_available_on(self, day_idx):
        return 0 <= day_idx < len(self.availability) and bool(self.availability[day_idx])


def normalize_cell(cell):
    if cell is None:
        return None
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip().replace('\u3000', ' ')
    if isinstance(v, (int, float)):
        return str(int(v))
    return str(v)


def read_students(input_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    male_list = []
    female_list = []

    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        is_male = '男' in sheetname

        # assume header in first row, data from row 2
        # 实际表格结构观察：
        # 列 A 可能为 '男生'/'女生' 标签（分组），列 B 为姓名，列 C 为时间（节次），列 D-H 为 周一..周五 的可用性
        # 所以我们按行读取，维护一个当前性别标记（当列 A 出现 '男生'/'女生' 时更新），姓名以列 B 为准
        current_is_male = None
        for r in range(2, sheet.max_row + 1):
            col_a = normalize_cell(sheet.cell(row=r, column=1))
            col_b = normalize_cell(sheet.cell(row=r, column=2))

            # 更新当前性别分组
            if col_a:
                a = col_a.replace(' ', '').replace('\u3000', '')
                if a == '男生' or a == '男':
                    current_is_male = True
                elif a == '女生' or a == '女':
                    current_is_male = False

            # 如果这一行没有姓名，跳过
            if not col_b:
                continue

            name = col_b
            # 读取周一到周五（列 D..H -> 4..8）
            availability = []
            for c in range(4, 4 + DAYS_PER_WEEK):
                val = normalize_cell(sheet.cell(row=r, column=c))
                ok = False
                if val is None:
                    ok = False
                else:
                    if str(val).strip() in ('是', 'Y', 'y', '1', 'yes'):
                        ok = True
                    else:
                        ok = str(val).strip() not in ('', '否', '0', 'no', 'n')
                availability.append(ok)

            # 如果当前性别未知，尝试从 sheet 名称判断
            if current_is_male is None:
                inferred = None
                if '男' in sheetname and '女' not in sheetname:
                    inferred = True
                elif '女' in sheetname and '男' not in sheetname:
                    inferred = False
                else:
                    inferred = True  # 默认男
                current_is_male = inferred

            student = Student(name, current_is_male, availability)
            if student.is_male:
                male_list.append(student)
            else:
                female_list.append(student)

    return male_list, female_list


def select_students(students, count, day_idx):
    selected = []
    # filter available and not exceeded weekly limit
    available = [s for s in students if s.is_available_on(day_idx) and s.weekly_count < WEEKLY_LIMIT]
    available.sort(key=lambda s: (s.weekly_count, s.name))

    for s in available:
        if len(selected) >= count:
            break
        if s.name in selected:
            continue
        selected.append(s.name)
        s.weekly_count += 1

    # if still not enough, relax weekly limit
    if len(selected) < count:
        remaining = [s for s in students if s.is_available_on(day_idx) and s.name not in selected]
        remaining.sort(key=lambda s: (s.weekly_count, s.name))
        for s in remaining:
            if len(selected) >= count:
                break
            selected.append(s.name)
            s.weekly_count += 1

    return selected


def generate_schedule(males, females):
    # reset counts
    for s in males + females:
        s.weekly_count = 0

    week_schedule = []
    for day_idx in range(DAYS_PER_WEEK):
        male_assign = select_students(males, MALE_PER_DAY, day_idx)
        female_assign = select_students(females, FEMALE_PER_DAY, day_idx)
        # ensure uniqueness within same day (male/female pools are disjoint by sheet classification)
        week_schedule.append((DAY_NAMES[day_idx], male_assign, female_assign))
    return week_schedule


def export_schedule(schedule, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '排班表'

    headers = ['日期', '男1', '男2', '男3', '男4', '女1', '女2']
    ws.append(headers)

    for day, males, females in schedule:
        row = [day]
        # ensure exactly 4 male cols
        row.extend(males + [''] * max(0, MALE_PER_DAY - len(males)))
        # exactly 2 female cols
        row.extend(females + [''] * max(0, FEMALE_PER_DAY - len(females)))
        ws.append(row)

    for col in range(1, 1 + len(headers)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].auto_size = True

    wb.save(output_path)
    print('导出完成:', output_path)


def main():
    if not os.path.exists(INPUT_PATH):
        raise SystemExit(f"未找到输入文件: {INPUT_PATH}")

    males, females = read_students(INPUT_PATH)
    print(f'读取完成：男生 {len(males)} 人，女生 {len(females)} 人')

    schedule = generate_schedule(males, females)
    for day, m, f in schedule:
        print(day, '男:', m, '女:', f)

    export_schedule(schedule, OUTPUT_PATH)


if __name__ == '__main__':
    main()

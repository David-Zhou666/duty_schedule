#!/usr/bin/env python3
import openpyxl, os
ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
path = os.path.join(ROOT, 'duty_schedule_python.xlsx')
if not os.path.exists(path):
    raise SystemExit('输出文件不存在: ' + path)
wb = openpyxl.load_workbook(path)
ws = wb.active
# 找到周二行
row_idx = None
for r in range(1, ws.max_row+1):
    if ws.cell(row=r, column=1).value == '周二':
        row_idx = r
        break
if row_idx is None:
    raise SystemExit('未找到 周二 行')
# 女生列 6,7
f1 = ws.cell(row=row_idx, column=6).value
f2 = ws.cell(row=row_idx, column=7).value
print('当前周二女生:', f1, f2)
# 如果第二个为空，尝试填入 潘玥函（来自原表候选）
candidate = '潘玥函'
if not f1:
    ws.cell(row=row_idx, column=6, value=candidate)
    print('填入 女1 ->', candidate)
elif not f2:
    ws.cell(row=row_idx, column=7, value=candidate)
    print('填入 女2 ->', candidate)
else:
    print('周二女生已满，不做修改')
wb.save(path)
print('保存完成:', path)

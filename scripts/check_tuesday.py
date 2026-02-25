#!/usr/bin/env python3
import openpyxl, os
p = r'src/main/resources/duty_temple.xlsx'
out = r'duty_schedule_python.xlsx'
DAYS_PER_WEEK = 5
DAY_NAMES = ['周一','周二','周三','周四','周五']

def normalize(x):
    if x is None: return None
    return str(x).strip()

wb = openpyxl.load_workbook(p, data_only=True)
print('sheets:', wb.sheetnames)

candidates = []
for sheetname in wb.sheetnames:
    sh = wb[sheetname]
    current_is_male = None
    for r in range(2, sh.max_row+1):
        a = normalize(sh.cell(row=r, column=1).value)
        b = normalize(sh.cell(row=r, column=2).value)
        if a:
            aa = a.replace(' ','').replace('\u3000','')
            if aa in ('男生','男'):
                current_is_male = True
            elif aa in ('女生','女'):
                current_is_male = False
        if not b:
            continue
        name = b
        # 周二在列5
        val = normalize(sh.cell(row=r, column=5).value)
        ok = False
        if val and val in ('是','Y','y','1','yes'):
            ok = True
        else:
            if val and str(val).strip() not in ('','否','0','no','n'):
                ok = True
        if current_is_male is False and ok:
            candidates.append((sheetname, name))

print('\n周二可用女生候选总数:', len(candidates))
for s,n in candidates:
    print('-', s, n)

# 读取生成的排班文件，查看周二已分配女生
if os.path.exists(out):
    wb2 = openpyxl.load_workbook(out, data_only=True)
    sh2 = wb2.active
    # find row for 周二
    assigned = []
    for r in range(1, sh2.max_row+1):
        cell = sh2.cell(row=r, column=1).value
        if cell == '周二':
            # female cols are columns 6 and 7 (1-based) per script
            f1 = sh2.cell(row=r, column=6).value
            f2 = sh2.cell(row=r, column=7).value
            assigned = [x for x in [f1,f2] if x]
            break
    print('\n周二已分配女生:', assigned)
else:
    print('\n未找到输出文件', out)

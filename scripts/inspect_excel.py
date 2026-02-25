#!/usr/bin/env python3
import openpyxl, os
p = r'C:\Users\Asus\CodeBuddy\20260221191406\src\main\resources\duty_temple.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
print('sheets:', wb.sheetnames)
for name in wb.sheetnames:
    print('\n---', name)
    sh = wb[name]
    max_r = min(10, sh.max_row)
    for r in range(1, max_r+1):
        vals = []
        for c in range(1, 1+8):
            cell = sh.cell(row=r, column=c).value
            vals.append(cell)
        print('r', r, vals)

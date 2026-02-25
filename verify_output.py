import openpyxl

file_path = 'duty_schedule_result.xlsx'
wb = openpyxl.load_workbook(file_path)

print("="*70)
print("   生成的排班表验证")
print("="*70)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    print("\n排班表内容:")
    print("-" * 70)
    print(f"{'周次':<8} | {'男生值班(4人)':<30} | {'女生值班(2人)':<30}")
    print("-" * 70)
    
    for row_idx, row in enumerate(ws.iter_rows(max_row=6, values_only=True), 1):
        if row_idx == 1:
            continue  # 跳过表头
        if row and row[0]:
            day = str(row[0])
            males = str(row[1]) if row[1] else "-"
            females = str(row[2]) if row[2] else "-"
            
            # 统计人数
            male_count = len([m for m in males.split('、') if m and m != '-'])
            female_count = len([f for f in females.split('、') if f and f != '-'])
            
            status_m = "✓" if male_count == 4 else f"⚠({male_count})"
            status_f = "✓" if female_count == 2 else f"⚠({female_count})"
            
            print(f"{day:<8} | {males:<30} {status_m} | {females:<30} {status_f}")

print("\n" + "="*70)
print("验证规则:")
print("✓ = 满足要求 (男生4人,女生2人)")  
print("⚠(n) = 不满足要求 (实际n人)")
print("="*70)

wb.close()

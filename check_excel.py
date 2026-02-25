import openpyxl

file_path = 'src/main/resources/duty_temple.xlsx'
wb = openpyxl.load_workbook(file_path)

print(f"Sheet数: {len(wb.sheetnames)}")
print(f"Sheet名称: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    print("\n数据内容（前15行）:")
    print("-" * 80)
    
    for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=False), 1):
        line = f"行{row_idx:2d}: "
        for cell in row:
            val = cell.value
            if val is not None:
                line += f"[{val}] "
        print(line)
    print()
